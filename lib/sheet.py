from typing import Optional

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SHEETNAME_MAXLEN = 31

def sanitize_sheet_name(name: str) -> str:
    # Excel sheet name restrictions
    invalid = set('[]:*?/\\')
    cleaned = ''.join(c if c not in invalid else '_' for c in name)
    if len(cleaned) > SHEETNAME_MAXLEN:
        cleaned = cleaned[:SHEETNAME_MAXLEN]
    return cleaned or "Sheet"


def apply_frozen_header(ws, headers, freeze_panes_cell: Optional[str] = "A2"):
    """Apply bold + gray header styling to row 1 across the given number of headers
    and optionally freeze panes at the specified cell (default A2).
    """
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if freeze_panes_cell:
        ws.freeze_panes = freeze_panes_cell


def apply_header_and_column_widths(ws, headers, column_widths=None, freeze_panes_cell: Optional[str] = "A2"):
    """Common helper to style header (row 1), freeze panes, and set column widths.
    - headers: list of header titles (used to know how many columns to style)
    - column_widths: either a list/tuple matching headers length, or a dict mapping
      column letters (e.g., 'A') to widths.
    - freeze_panes_cell: e.g., "A2" to freeze top row
    """
    # Apply header style and optional freeze
    apply_frozen_header(ws, headers, freeze_panes_cell)

    # Apply column widths, if provided
    if column_widths:
        if isinstance(column_widths, (list, tuple)):
            for idx, width in enumerate(column_widths, start=1):
                try:
                    if width is not None:
                        col_letter = get_column_letter(idx)
                        ws.column_dimensions[col_letter].width = width
                except Exception:
                    pass
        elif isinstance(column_widths, dict):
            for col_letter, width in column_widths.items():
                try:
                    if width is not None:
                        ws.column_dimensions[str(col_letter)].width = width
                except Exception:
                    pass


def apply_wrap_to_all_cells(ws):
    """Ensure wrap_text and top vertical alignment on all cells in the worksheet."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            # Preserve existing horizontal alignment if set
            horiz = getattr(cell.alignment, 'horizontal', None) if cell.alignment else None
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=horiz)
