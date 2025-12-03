import pandas as pd
import os
import argparse
import json
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# CẤU HÌNH (CONFIGURATION)
# ==========================================

# --- Bundle Info Config ---
BUNDLE_DEFAULTS = {
    "OLD": "bundle_info1.xlsx",
    "NEW": "bundle_info2.xlsx",
    "OUT": "bundle_info.xlsx",
    "REPORT": "report_bundle_info.txt"
}
BUNDLE_PATCH_HEADER = ["Bundle path suffix", "PathID", "Object selector", "Original", "Translated", "Notes"]
BUNDLE_PATCH_WIDTHS = [40, 16, 40, 60, 60, 60]
BUNDLE_INFO_WIDTHS = [40, 60, 20]

# --- Translate Config ---
TRANS_DEFAULTS = {
    "OLD": "translate1.xlsx",
    "NEW": "translate2.xlsx",
    "OUT": "translate.xlsx",
    "REPORT": "report_translate.txt"
}
TRANS_COMMON_HEADER = ["ID", "Original", "Chinese", "MTL", "Edited", "QA 1", "QA 2", "QA 3"]
TRANS_COL_WIDTHS = [32, 60, 60, 60, 60, 14, 14, 14]
TRANS_SYSTEM_SHEETS = ["Metadata", "Overview", "Knowledge base", "Summaries", "Patch addresses"]


# ==========================================
# CÁC HÀM HỖ TRỢ (HELPER FUNCTIONS)
# ==========================================

def clean_note_content(note_text):
    """Làm sạch cột Notes của Bundle Info (xóa tag cũ)"""
    if not isinstance(note_text, str): return ""
    lines = note_text.split('\n')
    clean_lines = []
    for line in lines:
        if line.strip().startswith('[') and ']:' in line: continue
        clean_lines.append(line)
    return "\n".join(clean_lines).strip()


def format_bundle_excel(file_path):
    """Định dạng chuẩn cho file Bundle Info"""
    print("    [-] Đang định dạng file Bundle Info...")
    try:
        wb = load_workbook(file_path)

        ws = None
        for s in wb.sheetnames:
            if s.lower() == "patch addresses": ws = wb[s]

        if ws:
            for col_idx, width in enumerate(BUNDLE_PATCH_WIDTHS, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=2).number_format = '@'
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        if "Bundle Info" in wb.sheetnames:
            ws = wb["Bundle Info"]
            for col_idx, width in enumerate(BUNDLE_INFO_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        wb.save(file_path)
    except Exception as e:
        print(f"    [!] Lỗi định dạng Bundle: {e}")


def format_translate_excel(file_path):
    """Định dạng chuẩn cho file Translate"""
    print("    [-] Đang định dạng file Translate...")
    try:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            if ws.title in TRANS_SYSTEM_SHEETS: continue

            for i, width in enumerate(TRANS_COL_WIDTHS):
                col_letter = get_column_letter(i + 1)
                ws.column_dimensions[col_letter].width = width

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            max_row = ws.max_row
            max_col = ws.max_column
            if max_row > 1:
                for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if cell.column == 1:
                            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        wb.save(file_path)
    except Exception as e:
        print(f"    [!] Lỗi định dạng Translate: {e}")


# ==========================================
# REPORTING
# ==========================================

def write_bundle_report(report_path, logs, stats):
    print(f"    [-] Đang xuất báo cáo Bundle: {report_path}")
    changed_logs = [l for l in logs if l['status'] == 'CONTENT_CHANGED']
    new_logs = [l for l in logs if l['status'] == 'NEW_UNMATCHED']

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=== BUNDLE INFO MERGE REPORT ===\n")
        f.write(f"Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Tổng số dòng: {stats['total']}\n")
        f.write("-" * 40 + "\n")
        f.write(f"1. Nội dung gốc thay đổi: {len(changed_logs)} (Cần kiểm tra)\n")
        f.write(f"2. Dòng mới chưa dịch:    {len(new_logs)}\n")
        f.write(f"3. Khớp hoàn hảo:         {stats.get('PERFECT_MATCH', 0)}\n")
        f.write("=" * 40 + "\n\n")

        if changed_logs:
            f.write(">>> DANH SÁCH THAY ĐỔI TEXT GỐC <<<\n\n")
            for item in changed_logs:
                f.write(f"[ID]: {item['path_id']} ({item['bundle']})\n")
                f.write(f"  Cũ: {item['original_old']}\n")
                f.write(f"  Mới: {item['original_new']}\n")
                f.write("-" * 30 + "\n")
            f.write("\n")

        if new_logs:
            f.write(">>> DANH SÁCH DÒNG MỚI <<<\n\n")
            for item in new_logs:
                f.write(f"[ID]: {item['path_id']} ({item['bundle']})\n")
                f.write(f"  Gốc: {item['original_new']}\n")


def write_translate_report(report_path, logs):
    print(f"    [-] Đang xuất báo cáo Translate: {report_path}")
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=== TRANSLATE EXCEL MERGE REPORT ===\n")
        f.write(f"Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 40 + "\n\n")

        if not logs:
            f.write("Không phát hiện thay đổi nào trong Text gốc.\n")
        else:
            f.write(f"Tìm thấy {len(logs)} thay đổi trong Text gốc.\n\n")
            for item in logs:
                f.write(f"[Sheet]: {item['sheet']} | [ID]: {item['id']}\n")
                f.write(f"  Cột: {item['field']}\n")
                f.write(f"  Cũ: {item['old']}\n")
                f.write(f"  Mới: {item['new']}\n")
                f.write("-" * 30 + "\n")


# ==========================================
# MODULE 1: BUNDLE INFO MERGE
# ==========================================

def run_bundle_merge(old_file, new_file, out_file, report_file):
    print(f"\n[1/2] ĐANG CHẠY BUNDLE INFO MERGE...")
    if not os.path.exists(old_file) or not os.path.exists(new_file):
        print(f"    [!] Bỏ qua: Không tìm thấy file {old_file} hoặc {new_file}")
        return

    try:
        df_old = pd.read_excel(old_file, sheet_name="Patch addresses", dtype={'PathID': str})
        df_new = pd.read_excel(new_file, sheet_name="Patch addresses", dtype={'PathID': str})
        df_info_new = pd.read_excel(new_file, sheet_name="Bundle Info", dtype=str)
    except Exception as e:
        print(f"    [!] Lỗi đọc file: {e}")
        return

    df_old.columns = [c.strip() for c in df_old.columns]
    df_new.columns = [c.strip() for c in df_new.columns]
    for col in BUNDLE_PATCH_HEADER:
        if col not in df_new.columns: df_new[col] = ""

    df_new['Translated'] = df_new['Translated'].astype(object)
    df_new['Notes'] = df_new['Notes'].astype(object)

    # Indexing
    context_map = {}
    id_map = {}
    for index, row in df_old.iterrows():
        suffix = str(row.get('Bundle path suffix', '')).strip()
        path_id = str(row.get('PathID', '')).strip()
        selector = str(row.get('Object selector', '')).strip()
        original = str(row.get('Original', '')).strip()
        raw_notes = str(row.get('Notes', ''))
        clean_notes = clean_note_content(raw_notes)
        trans = row.get('Translated', '')

        if pd.notna(trans) and str(trans).lower() != 'nan' and str(trans).strip() != '':
            key_ctx = (suffix, original, clean_notes)
            context_map[key_ctx] = trans
            key_id = (suffix, path_id, selector)
            id_map[key_id] = {'trans': trans, 'orig_old': original}

    # Merging
    logs = []
    for index, row in df_new.iterrows():
        suffix = str(row.get('Bundle path suffix', '')).strip()
        path_id = str(row.get('PathID', '')).strip()
        selector = str(row.get('Object selector', '')).strip()
        orig_new = str(row.get('Original', '')).strip()
        notes_new = str(row.get('Notes', '')).strip()

        key_ctx = (suffix, orig_new, notes_new)
        key_id = (suffix, path_id, selector)

        found_trans = None
        status = ""
        orig_old_log = None

        if key_ctx in context_map:
            found_trans = context_map[key_ctx]
            status = "PERFECT_MATCH"
        elif key_id in id_map:
            old_data = id_map[key_id]
            found_trans = old_data['trans']
            orig_old_log = old_data['orig_old']
            status = "CONTENT_CHANGED"
        else:
            status = "NEW_UNMATCHED"

        if found_trans:
            df_new.at[index, 'Translated'] = found_trans

        logs.append({
            "status": status,
            "path_id": path_id,
            "bundle": suffix,
            "original_new": orig_new,
            "original_old": orig_old_log
        })

    # Saving
    df_new = df_new[BUNDLE_PATCH_HEADER]
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df_info_new.to_excel(writer, sheet_name="Bundle Info", index=False)
        df_new.to_excel(writer, sheet_name="Patch addresses", index=False)

    format_bundle_excel(out_file)
    stats = pd.DataFrame(logs)['status'].value_counts().to_dict()
    stats['total'] = len(df_new)
    write_bundle_report(report_file, logs, stats)
    print(f"    [v] Hoàn tất! Output: {out_file}")


# ==========================================
# MODULE 2: TRANSLATE MERGE
# ==========================================

def normalize_trans_columns(df):
    for col in TRANS_COMMON_HEADER:
        if col not in df.columns: df[col] = ""
    return df[TRANS_COMMON_HEADER]


def merge_trans_sheet(sheet_name, df_old, df_new, diff_records):
    df_old.columns = [str(c).strip() for c in df_old.columns]
    df_new.columns = [str(c).strip() for c in df_new.columns]

    if 'ID' not in df_old.columns or 'ID' not in df_new.columns: return df_new

    df_old['ID'] = df_old['ID'].astype(str)
    df_new['ID'] = df_new['ID'].astype(str)
    df_old = normalize_trans_columns(df_old)
    df_new = normalize_trans_columns(df_new)

    merged = pd.merge(df_old, df_new, on='ID', how='outer', suffixes=('_old', '_new'), indicator=True)
    final_rows = []

    for index, row in merged.iterrows():
        row_id = row['ID']
        merge_status = row['_merge']
        new_row = {'ID': row_id}

        orig_old = str(row['Original_old']) if pd.notna(row['Original_old']) else ""
        orig_new = str(row['Original_new']) if pd.notna(row['Original_new']) else ""
        chn_old = str(row['Chinese_old']) if pd.notna(row['Chinese_old']) else ""
        chn_new = str(row['Chinese_new']) if pd.notna(row['Chinese_new']) else ""

        if merge_status == 'left_only':
            new_row['Original'] = orig_old
            new_row['Chinese'] = chn_old
        else:
            new_row['Original'] = orig_new
            new_row['Chinese'] = chn_new
            if merge_status == 'both':
                if orig_old != orig_new:
                    diff_records.append(
                        {"sheet": sheet_name, "id": row_id, "field": "Original", "old": orig_old, "new": orig_new})
                if chn_old != chn_new:
                    diff_records.append(
                        {"sheet": sheet_name, "id": row_id, "field": "Chinese", "old": chn_old, "new": chn_new})

        for col in ["MTL", "Edited", "QA 1", "QA 2", "QA 3"]:
            val_old = row[f'{col}_old']
            val_new = row[f'{col}_new']
            if merge_status == 'right_only':
                new_row[col] = val_new if pd.notna(val_new) else ""
            else:
                new_row[col] = val_old if pd.notna(val_old) else ""

        final_rows.append(new_row)
    return pd.DataFrame(final_rows, columns=TRANS_COMMON_HEADER)


def run_translate_merge(old_file, new_file, out_file, report_file):
    print(f"\n[2/2] ĐANG CHẠY TRANSLATE MERGE...")
    if not os.path.exists(old_file) or not os.path.exists(new_file):
        print(f"    [!] Bỏ qua: Không tìm thấy file {old_file} hoặc {new_file}")
        return

    try:
        xls_old = pd.read_excel(old_file, sheet_name=None, engine='openpyxl', dtype={'ID': str})
        xls_new = pd.read_excel(new_file, sheet_name=None, engine='openpyxl', dtype={'ID': str})
    except Exception as e:
        print(f"    [!] Lỗi đọc file: {e}")
        return

    diff_records = []
    final_sheet_order = list(xls_old.keys())
    for sheet in xls_new.keys():
        if sheet not in final_sheet_order: final_sheet_order.append(sheet)

    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        for sheet_name in final_sheet_order:
            if sheet_name in TRANS_SYSTEM_SHEETS:
                if sheet_name in ["Metadata", "Patch addresses"]:
                    target_df = xls_new.get(sheet_name, xls_old.get(sheet_name))
                else:
                    target_df = xls_old.get(sheet_name, xls_new.get(sheet_name))
                if target_df is not None: target_df.to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            df_old = xls_old.get(sheet_name, pd.DataFrame(columns=TRANS_COMMON_HEADER))
            df_new = xls_new.get(sheet_name, pd.DataFrame(columns=TRANS_COMMON_HEADER))

            if sheet_name not in xls_new:
                df_old.to_excel(writer, sheet_name=sheet_name, index=False)
                continue
            if sheet_name not in xls_old:
                df_new = normalize_trans_columns(df_new)
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            final_df = merge_trans_sheet(sheet_name, df_old, df_new, diff_records)
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    format_translate_excel(out_file)
    write_translate_report(report_file, diff_records)
    print(f"    [v] Hoàn tất! Output: {out_file}")


# ==========================================
# MAIN
# ==========================================

def main():
    parser = argparse.ArgumentParser(description="Universal Auto Merge Tool")

    # Bundle Args
    parser.add_argument('--bundle-old', default=BUNDLE_DEFAULTS["OLD"])
    parser.add_argument('--bundle-new', default=BUNDLE_DEFAULTS["NEW"])
    parser.add_argument('--bundle-out', default=BUNDLE_DEFAULTS["OUT"])
    parser.add_argument('--bundle-report', default=BUNDLE_DEFAULTS["REPORT"])

    # Translate Args
    parser.add_argument('--trans-old', default=TRANS_DEFAULTS["OLD"])
    parser.add_argument('--trans-new', default=TRANS_DEFAULTS["NEW"])
    parser.add_argument('--trans-out', default=TRANS_DEFAULTS["OUT"])
    parser.add_argument('--trans-report', default=TRANS_DEFAULTS["REPORT"])

    args = parser.parse_args()

    # Tự động chạy cả 2 nếu file tồn tại
    print("=== UNIVERSAL MERGE TOOL STARTING ===")

    run_bundle_merge(args.bundle_old, args.bundle_new, args.bundle_out, args.bundle_report)
    run_translate_merge(args.trans_old, args.trans_new, args.trans_out, args.trans_report)

    print("\n=== ALL TASKS FINISHED ===")


if __name__ == "__main__":
    main()