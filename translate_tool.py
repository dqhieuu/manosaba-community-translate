import os
import re
import shutil
import sys
import traceback
from typing import List, Tuple, Optional
from pathlib import Path
from collections import defaultdict

from PIL import Image
from openai import OpenAI
from openai.types.shared_params import Reasoning
from openpyxl import Workbook, load_workbook
import UnityPy
import yaml

from lib.bin import validate_bin_patch_map
from lib.steam import get_steam_game_path
from lib.tree_traversal import set_by_selector
from lib.utils import is_file_editable, is_running_in_exe
from lib.text import is_alnum_start, trim_blank_lines
from lib.sheet import sanitize_sheet_name, apply_header_and_column_widths, apply_wrap_to_all_cells

IGNORED_BUNDLE_SUFFIXES = ['general-managedtext_assets_all.bundle']

ROOT = sys._MEIPASS if is_running_in_exe() else os.path.dirname(os.path.abspath(__file__))
ORIGINAL_DIR = os.path.join(ROOT, "original")
TRANSLATED_DIR = os.path.join(ROOT, "translated")
PATCHES_DIR = os.path.join(ROOT, "patches")

XLSX_PATH = os.path.join(ROOT, "translate.xlsx")
ADDRESSES_PATH = os.path.join(PATCHES_DIR, "addresses.txt")

KNOWLEDGE_SHEETNAME = "Knowledge base"
METADATA_SHEETNAME = "Metadata"
OVERVIEW_SHEETNAME = "Overview"
SUMMARIES_SHEETNAME = "Summaries"
PATCH_SHEETNAME = "Patch addresses"

SPECIAL_SHEETS = {OVERVIEW_SHEETNAME, METADATA_SHEETNAME, KNOWLEDGE_SHEETNAME, SUMMARIES_SHEETNAME, PATCH_SHEETNAME}

COMMON_TRANSLATE_HEADER = ["ID", "Original", "Chinese", "MTL", "Edited"]
METADATA_HEADER = ["Sheet name", "Mapped file name", "File type"]
KNOWLEDGE_HEADER = ["Knowledge"]
OVERVIEW_HEADER = ["Act", "Chapter", "File", "Total Lines", "MTL %", "Edited %"]
SUMMARIES_HEADER = ["Sheet name", "Summary"]
PATCH_HEADER = ["Bundle path suffix", "PathID", "Object selector", "Original", "Translated", "Notes"]

INITIAL_PROJECT_HEADER = [
    "Project type: Visual Novel translation.",
    "Goal: Produce high‑quality, natural translations suitable for a visual novel UI/dialogue.",
    "Guidelines: Preserve placeholders, variables, control codes, and line breaks. Maintain speaker tone, honorifics where appropriate, and context.",
    "Text between <ruby> should be converted to Romaji"
]


def ensure_patch_sheet(wb):
    if PATCH_SHEETNAME not in wb.sheetnames:
        ws = wb.create_sheet(title=PATCH_SHEETNAME)
        ws.append(PATCH_HEADER)
        # Style, freeze, and set column widths via common helper
        apply_header_and_column_widths(ws, PATCH_HEADER, [50, 16, 60, 60, 60, 60])
        apply_wrap_to_all_cells(ws)
    else:
        ws = wb[PATCH_SHEETNAME]
    # Enforce PathID column (B) as plain text
    enforce_patch_pathid_text(ws)
    return ws

def load_patches_from_files() -> dict:
    """Load patch addresses from multiple sources and merge them.
    Priority (last one wins on duplicates):
      1) patches/addresses.txt (YAML)
      2) translate.xlsx -> "Patch addresses" sheet
      3) bundle_info.xlsx -> "Patch Addresses" sheet
    Schema returned: { suffix: { path_id(str): [ {object_selector: str, patched_value: any}, ... ] } }
    """
    merged: dict = {}

    def _merge_entry(store: dict, suffix: str, pid: str, selector: str, value):
        if not suffix or not pid or not selector:
            return
        id_map = store.setdefault(suffix, {})
        entries = id_map.setdefault(str(pid), [])
        # Replace if selector already exists, else append
        for ent in entries:
            if ent.get('object_selector') == selector:
                ent['patched_value'] = value
                return
        entries.append({'object_selector': selector, 'patched_value': value})

    # 1) YAML file
    if os.path.exists(ADDRESSES_PATH):
        try:
            with open(ADDRESSES_PATH, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    data = yaml.safe_load(content) or {}
                    # normalize into merged
                    if isinstance(data, dict):
                        for suf, id_map in data.items():
                            if not isinstance(id_map, dict):
                                continue
                            for pid, entries in id_map.items():
                                try:
                                    pid_str = str(pid)
                                except Exception:
                                    pid_str = str(pid)
                                if isinstance(entries, list):
                                    for ent in entries:
                                        if not isinstance(ent, dict):
                                            continue
                                        selector = ent.get('object_selector')
                                        value = ent.get('patched_value')
                                        if selector is not None:
                                            _merge_entry(merged, suf, pid_str, selector, value)
        except Exception as e:
            print(f"Warning: Failed to read {ADDRESSES_PATH}: {e}")

    # Helper to read a patch sheet from a workbook path
    def _gather_from_workbook(xlsx_path: str):
        if not os.path.exists(xlsx_path):
            return
        try:
            wb = load_workbook(xlsx_path)
        except Exception:
            return
        # Find a sheet whose name matches 'Patch addresses' (case-insensitive, with or without capital A)
        target_ws = None
        for s in wb.sheetnames:
            if s.lower().strip() in { 'patch addresses' }:
                target_ws = wb[s]
                break
        if target_ws is None:
            return
        # Map headers
        headers = [ (target_ws.cell(row=1, column=c).value or '').strip() for c in range(1, target_ws.max_column + 1) ]
        name_to_idx = { (h or '').strip().lower(): i+1 for i, h in enumerate(headers) }
        needed = ['bundle path suffix', 'pathid', 'object selector']
        if not all(col in name_to_idx for col in needed):
            return
        col_suffix = name_to_idx['bundle path suffix']
        col_pathid = name_to_idx['pathid']
        col_selector = name_to_idx['object selector']
        col_original = name_to_idx.get('original')
        col_translated = name_to_idx.get('translated')
        max_row = target_ws.max_row or 1
        for r in range(2, max_row + 1):
            suf = (target_ws.cell(row=r, column=col_suffix).value or '').strip()
            pid = str((target_ws.cell(row=r, column=col_pathid).value or '').strip())
            selector = (target_ws.cell(row=r, column=col_selector).value or '').strip()
            if not suf or not pid or not selector:
                continue
            val_t = (target_ws.cell(row=r, column=col_translated).value if col_translated else None)
            val_o = (target_ws.cell(row=r, column=col_original).value if col_original else None)
            val_t = (str(val_t) if val_t is not None else '').strip()
            val_o = (str(val_o) if val_o is not None else '').strip()
            value = val_t if val_t != '' else val_o
            if value == '':
                continue
            _merge_entry(merged, suf, pid, selector, value)

    # 2) translate.xlsx
    _gather_from_workbook(XLSX_PATH)
    # 3) bundle_info.xlsx (same ROOT)
    _gather_from_workbook(os.path.join(ROOT, 'bundle_info.xlsx'))

    return merged

def populate_patch_sheet_from_file(wb, update_instead_of_overwrite: bool = True) -> None:
    ws = ensure_patch_sheet(wb)
    has_rows = ws.max_row and ws.max_row > 1
    data = load_patches_from_files()
    if not data:
        return

    # Clear existing rows if not only_if_empty
    if not update_instead_of_overwrite and has_rows:
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)
        has_rows = False

    if not has_rows:
        # Sheet empty -> populate all entries from file (original behavior)
        for bundle_suffix, id_map in data.items():
            for path_id, entries in id_map.items():
                pid_str = str(path_id)
                for ent in entries:
                    selector = ent.get('object_selector', '')
                    val = ent.get('patched_value', '')
                    ws.append([bundle_suffix, pid_str, selector, val, "", ""])  # Original=val, Translated empty; Notes empty
    else:
        # Sheet has existing rows and only_if_empty=True -> merge: add missing and fill blanks
        # Build index of existing rows: key -> row number
        headers = [(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        try:
            col_suffix = headers.index("Bundle path suffix") + 1
            col_pathid = headers.index("PathID") + 1
            col_selector = headers.index("Object selector") + 1
            col_original = headers.index("Original") + 1
            col_translated = headers.index("Translated") + 1
        except ValueError:
            print("Unmatched headers in Patch sheet. Skipping updating sheet from file.")
        else:
            index = {}
            for r in range(2, ws.max_row + 1):
                suffix = (ws.cell(row=r, column=col_suffix).value or "").strip()
                pid = str((ws.cell(row=r, column=col_pathid).value or "").strip())
                selector = (ws.cell(row=r, column=col_selector).value or "").strip()
                if suffix and pid and selector:
                    index[(suffix, pid, selector)] = r

            # Merge from JSON
            for bundle_suffix, id_map in data.items():
                for path_id, entries in id_map.items():
                    pid_str = str(path_id)
                    for ent in entries:
                        selector = ent.get('object_selector', '')
                        val = ent.get('patched_value', '')
                        key = (bundle_suffix, pid_str, selector)
                        r = index.get(key)
                        if r is None:
                            ws.append([bundle_suffix, pid_str, selector, val, ""])  # New line
                        else:
                            # Update cells if empty
                            orig_cell = ws.cell(row=r, column=col_original)
                            trans_cell = ws.cell(row=r, column=col_translated)
                            if (orig_cell.value is None) or (str(orig_cell.value).strip() == ""):
                                orig_cell.value = val
                            if (trans_cell.value is None) or (str(trans_cell.value).strip() == ""):
                                trans_cell.value = val
    apply_wrap_to_all_cells(ws)
    # Enforce PathID text format after population/merge
    enforce_patch_pathid_text(ws)

def dump_patches_from_files() -> None:
    """Dump merged patches from all sources into patches/addresses.txt.
    Uses load_patches_from_files to aggregate from YAML + translate.xlsx + bundle_info.xlsx.
    """
    merged = load_patches_from_files() or {}
    os.makedirs(PATCHES_DIR, exist_ok=True)
    try:
        with open(ADDRESSES_PATH, 'w', encoding='utf-8') as f:
            yaml.safe_dump(merged, f, allow_unicode=True, sort_keys=True)
        print(f"Wrote merged patches to {ADDRESSES_PATH} ({sum(len(v) for v in merged.values())} path groups)")
    except Exception as e:
        print(f"Error writing {ADDRESSES_PATH}: {e}")

def enforce_patch_pathid_text(ws):
    """Ensure PathID column (B) is plain text in Excel and values are stored as strings."""
    try:
        max_row = ws.max_row or 1
    except Exception:
        return
    for r in range(1, max_row + 1):
        cell = ws.cell(row=r, column=2)  # Column B
        # Always set number format to text
        cell.number_format = "@"
        # Convert data rows to string values (leave header as-is)
        if r == 1:
            continue
        if cell.value is not None:
            cell.value = str(cell.value).strip()


def parse_type2(lines: List[str]) -> List[Tuple[str, str, str]]:
    """Return list of (ID, original, localized)"""
    results = []
    last_comment_block: List[str] = []

    for idx, raw in enumerate(lines):
        line = raw.rstrip('\n')
        if not line.strip():
            continue
        if line.lstrip().startswith(';'):
            if idx == 0: continue # Ignore the first metadata line
            comment = line.lstrip()[1:]
            if comment.startswith(' '):
                comment = comment[1:]
            last_comment_block.append(comment)
            continue
        m = re.match(r"^\s*([^:]+):\s*(.*)$", line)
        if m:
            _id = m.group(1).strip()
            localized = m.group(2)
            original = "\n".join(last_comment_block).strip()
            localized = trim_blank_lines(localized)
            results.append((_id, original, localized))
            last_comment_block = []
        else:
            _id = line.strip()
            results.append((_id, "\n".join(last_comment_block).strip(), ""))
            last_comment_block = []
    return results

def parse_type1(lines: List[str]) -> List[Tuple[str, str, str]]:
    """Parse blocks of:
    # ID
    ; Original (one or more comment lines)
    Localized (one or more non-comment lines until next '#')
    Return list of (ID, original, localized).
    """
    results: List[Tuple[str, str, str]] = []
    i = 0
    n = len(lines)
    while i < n:
        s = lines[i].rstrip('\n')
        if not s.strip():
            i += 1
            continue
        if s.lstrip().startswith('#'):
            id_part = s.lstrip()[1:]
            if id_part.startswith(' '):
                id_part = id_part[1:]
            _id = id_part.strip()
            i += 1
            orig_lines: List[str] = []
            while i < n:
                t = lines[i].rstrip('\n')
                if t.lstrip().startswith(';'):
                    c = t.lstrip()[1:]
                    if c.startswith(' '):
                        c = c[1:]
                    orig_lines.append(c)
                    i += 1
                else:
                    break
            loc_lines: List[str] = []
            while i < n:
                t = lines[i].rstrip('\n')
                if t.lstrip().startswith('#'):
                    break
                if not t.lstrip().startswith(';'):
                    loc_lines.append(t)
                i += 1
            original = trim_blank_lines("\n".join(orig_lines))
            localized = trim_blank_lines("\n".join(loc_lines))
            results.append((_id, original, localized))
            continue
        else:
            i += 1
            continue
    return results

def detect_file_type(lines: List[str]) -> Optional[int]:
    for raw in lines:
        s = raw.rstrip('\n')
        if not s.strip():
            continue
        if s.lstrip().startswith(';'):
            continue
        if s.lstrip().startswith('#'):
            return 1
        if is_alnum_start(s):
            return 2
    return None

def get_content_sheets(wb):
    return [s for s in wb.sheetnames if s not in SPECIAL_SHEETS]

def update_overview(wb):
    if OVERVIEW_SHEETNAME not in wb.sheetnames:
        ov_ws = wb.create_sheet(title=OVERVIEW_SHEETNAME, index=0)
        header = OVERVIEW_HEADER
        ov_ws.append(header)
        apply_header_and_column_widths(ov_ws, header, [20, 20, 40, 20, 20, 20])
        apply_wrap_to_all_cells(ov_ws)
    else:
        ov_ws = wb[OVERVIEW_SHEETNAME]
        for row in range(ov_ws.max_row, 1, -1):
            ov_ws.delete_rows(row)

    content_sheets = get_content_sheets(wb)
    structure = defaultdict(lambda: defaultdict(list))
    for sheet_name in content_sheets:
        if sheet_name.lower().startswith('common'):
            structure['Common']['Common'].append(sheet_name)
        else:
            match = re.match(r"^(Act\d+)_Chapter(\d+)_(.+)$", sheet_name)
            if not match:
                continue
            act, chapter, file_type = match.groups()
            structure[act][chapter].append(sheet_name)

    total_all_lines = 0
    total_all_mtl = 0
    total_all_edited = 0

    last_act = None
    last_chapter = None
    for act in sorted(structure.keys()):
        act_total_lines = 0
        act_mtl_completed = 0
        act_edited_completed = 0
        for chapter in sorted(structure[act].keys()):
            chapter_total_lines = 0
            chapter_mtl_completed = 0
            chapter_edited_completed = 0
            adv_total = trial_total = bad_total = common_total = 0
            adv_mtl = trial_mtl = bad_mtl = common_mtl = 0
            adv_edited = trial_edited = bad_edited = common_edited = 0

            for sheet_name in sorted(structure[act][chapter]):
                ws = wb[sheet_name]
                headers = [(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
                try:
                    col_id = headers.index("ID") + 1
                    col_mtl = headers.index("MTL") + 1
                    col_edited = headers.index("Edited") + 1
                except ValueError:
                    continue
                total_lines = 0
                mtl_completed = 0
                edited_completed = 0
                for r in range(2, ws.max_row + 1):
                    row_id = (ws.cell(row=r, column=col_id).value or "").strip()
                    if not row_id:
                        continue
                    total_lines += 1
                    mtl = (ws.cell(row=r, column=col_mtl).value or "").strip()
                    edited = (ws.cell(row=r, column=col_edited).value or "").strip()
                    if mtl:
                        mtl_completed += 1
                    if edited:
                        edited_completed += 1
                chapter_total_lines += total_lines
                chapter_mtl_completed += mtl_completed
                chapter_edited_completed += edited_completed
                act_total_lines += total_lines
                act_mtl_completed += mtl_completed
                act_edited_completed += edited_completed
                total_all_lines += total_lines
                total_all_mtl += mtl_completed
                total_all_edited += edited_completed

                if 'adv' in sheet_name.lower():
                    adv_total += total_lines
                    adv_mtl += mtl_completed
                    adv_edited += edited_completed
                elif 'trial' in sheet_name.lower():
                    trial_total += total_lines
                    trial_mtl += mtl_completed
                    trial_edited += edited_completed
                elif 'bad' in sheet_name.lower():
                    bad_total += total_lines
                    bad_mtl += mtl_completed
                    bad_edited += edited_completed
                elif 'common' in sheet_name.lower():
                    common_total += total_lines
                    common_mtl += mtl_completed
                    common_edited += edited_completed

                act_display = act if act != last_act else ""
                chapter_display = f"Chapter{chapter}" if chapter != last_chapter or act != last_act else ""
                mtl_perc = (mtl_completed / total_lines * 100) if total_lines > 0 else 0
                edited_perc = (edited_completed / total_lines * 100) if total_lines > 0 else 0
                ov_ws.append([
                    act_display,
                    chapter_display,
                    sheet_name,
                    total_lines,
                    f"{mtl_perc:.2f}%",
                    f"{edited_perc:.2f}%"
                ])
                last_act = act
                last_chapter = chapter

            # Add per-file-type totals for the chapter
            if adv_total > 0:
                adv_mtl_perc = (adv_mtl / adv_total * 100) if adv_total > 0 else 0
                adv_edited_perc = (adv_edited / adv_total * 100) if adv_total > 0 else 0
                ov_ws.append([
                    "" if act == last_act else act,
                    "" if chapter == last_chapter and act == last_act else f"Chapter{chapter}",
                    "Adv Total",
                    adv_total,
                    f"{adv_mtl_perc:.2f}%",
                    f"{adv_edited_perc:.2f}%"
                ])
            if trial_total > 0:
                trial_mtl_perc = (trial_mtl / trial_total * 100) if trial_total > 0 else 0
                trial_edited_perc = (trial_edited / trial_total * 100) if trial_total > 0 else 0
                ov_ws.append([
                    "" if act == last_act else act,
                    "" if chapter == last_chapter and act == last_act else f"Chapter{chapter}",
                    "Trial Total",
                    trial_total,
                    f"{trial_mtl_perc:.2f}%",
                    f"{trial_edited_perc:.2f}%"
                ])
            if bad_total > 0:
                bad_mtl_perc = (bad_mtl / bad_total * 100) if bad_total > 0 else 0
                bad_edited_perc = (bad_edited / bad_total * 100) if bad_total > 0 else 0
                ov_ws.append([
                    "" if act == last_act else act,
                    "" if chapter == last_chapter and act == last_act else f"Chapter{chapter}",
                    "Bad Total",
                    bad_total,
                    f"{bad_mtl_perc:.2f}%",
                    f"{bad_edited_perc:.2f}%"
                ])
            if common_total > 0:
                common_mtl_perc = (common_mtl / common_total * 100) if common_total > 0 else 0
                common_edited_perc = (common_edited / common_total * 100) if common_total > 0 else 0
                ov_ws.append([
                    "" if act == last_act else act,
                    "" if chapter == last_chapter and act == last_act else f"Chapter{chapter}",
                    "Common Total",
                    common_total,
                    f"{common_mtl_perc:.2f}%",
                    f"{common_edited_perc:.2f}%"
                ])

            # Add chapter total
            if chapter_total_lines > 0:
                chapter_mtl_perc = (chapter_mtl_completed / chapter_total_lines * 100) if chapter_total_lines > 0 else 0
                chapter_edited_perc = (chapter_edited_completed / chapter_total_lines * 100) if chapter_total_lines > 0 else 0
                ov_ws.append([
                    "" if act == last_act else act,
                    f"Chapter{chapter} Total",
                    "",
                    chapter_total_lines,
                    f"{chapter_mtl_perc:.2f}%",
                    f"{chapter_edited_perc:.2f}%"
                ])

        # Add act total
        if act_total_lines > 0:
            act_mtl_perc = (act_mtl_completed / act_total_lines * 100) if act_total_lines > 0 else 0
            act_edited_perc = (act_edited_completed / act_total_lines * 100) if act_total_lines > 0 else 0
            ov_ws.append([
                f"{act} Total",
                "",
                "",
                act_total_lines,
                f"{act_mtl_perc:.2f}%",
                f"{act_edited_perc:.2f}%"
            ])

    # Add grand total including Common
    if total_all_lines > 0:
        total_mtl_perc = (total_all_mtl / total_all_lines * 100) if total_all_lines > 0 else 0
        total_edited_perc = (total_all_edited / total_all_lines * 100) if total_all_lines > 0 else 0
        ov_ws.append([
            "Grand Total",
            "",
            "",
            total_all_lines,
            f"{total_mtl_perc:.2f}%",
            f"{total_edited_perc:.2f}%"
        ])

def _load_and_parse_original_txt(path: str):
    """Load a .txt file with utf-8.
    Returns tuple (ftype, data) where data is a list of (ID, Original, Localized).
    Returns (None, None) if type cannot be detected.
    """
    try:
        with open(path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except UnicodeDecodeError:
        print(f"File {path} cannot be decoded.")
        return None, None

    ftype = detect_file_type(lines)
    if ftype is None:
        return None, None

    if ftype == 1:
        data = parse_type1(lines)
    elif ftype == 2:
        data = parse_type2(lines)
    else:
        data = None
    return ftype, data

def _add_sheet_with_parsed_data(wb, base_sheet_name: str, data: List[Tuple[str, str, str]]):
    """Create a new sheet for the given data, avoiding name collisions.
    Returns the final sheet name used.
    """
    sheet_name = sanitize_sheet_name(base_sheet_name)
    existing = set(wb.sheetnames)
    if sheet_name in existing:
        suffix = 1
        while f"{sheet_name}_{suffix}" in existing:
            suffix += 1
        sheet_name = sanitize_sheet_name(f"{sheet_name}_{suffix}")
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COMMON_TRANSLATE_HEADER)
    apply_header_and_column_widths(ws, COMMON_TRANSLATE_HEADER, [32, 60, 60, 60, 60])
    for _id, original, localized in data:
        ws.append([_id, original, trim_blank_lines(localized), "", ""])
    apply_wrap_to_all_cells(ws)
    return sheet_name

def parse_original_files() -> None:
    if os.path.exists(XLSX_PATH):
        print(f"translate.xlsx already exists at {XLSX_PATH}. Skipping parse.")
        return

    if not os.path.isdir(ORIGINAL_DIR):
        print(f"Original directory not found: {ORIGINAL_DIR}")
        sys.exit(1)

    wb = Workbook()

    metadata_rows = []
    for fname in sorted(os.listdir(ORIGINAL_DIR)):
        if not fname.lower().endswith('.txt'):
            continue
        path = os.path.join(ORIGINAL_DIR, fname)
        ftype, data = _load_and_parse_original_txt(path)
        if ftype is None:
            print(f"Warning: Could not detect file type for {fname}. Skipping.")
            continue
        base_sheet_name = os.path.splitext(fname)[0]
        sheet_name = _add_sheet_with_parsed_data(wb, base_sheet_name, data)
        metadata_rows.append([sheet_name, fname, ftype])

    meta_ws = wb.create_sheet(title=METADATA_SHEETNAME)
    meta_ws.append(METADATA_HEADER)
    apply_header_and_column_widths(meta_ws, METADATA_HEADER, [32, 60, 12])

    for row in metadata_rows:
        meta_ws.append(row)

    apply_wrap_to_all_cells(meta_ws)

    kb_ws = wb.create_sheet(title=KNOWLEDGE_SHEETNAME)
    kb_ws.append(KNOWLEDGE_HEADER)
    apply_header_and_column_widths(kb_ws, KNOWLEDGE_HEADER, [100])
    for line in INITIAL_PROJECT_HEADER:
        kb_ws.append([line])
    apply_wrap_to_all_cells(kb_ws)

    sum_ws = wb.create_sheet(title=SUMMARIES_SHEETNAME)
    _sum_headers = SUMMARIES_HEADER
    sum_ws.append(_sum_headers)
    # Keep Summaries unfrozen as before
    apply_header_and_column_widths(sum_ws, _sum_headers, [32, 100])
    apply_wrap_to_all_cells(sum_ws)

    # Create Patch addresses sheet and populate from existing file if available
    ensure_patch_sheet(wb)

    update_overview(wb)

    wb.save(XLSX_PATH)
    print(f"translate.xlsx created at {XLSX_PATH}")

def get_knowledge_text(wb) -> str:
    if KNOWLEDGE_SHEETNAME in wb.sheetnames:
        ws = wb[KNOWLEDGE_SHEETNAME]
        parts = []
        for r in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if not r:
                continue
            v = (r[0] or "").strip()
            if v:
                parts.append(v)
        return "\n\n".join(parts).strip()
    return "\n\n".join(INITIAL_PROJECT_HEADER)

def generate_file_summary(client, sheet_name: str, rows: List[Tuple[str, str, str]]) -> str:
    context = "\n\n".join(
        f"ID: {row[0]}\nOriginal: {row[1] or '<empty>'}\nChinese: {row[2] or '<empty>'}"
        for row in rows
    )
    sys_prompt = (
        "You are a translator for a visual novel. Summarize the content of the following file in 2-3 concise lines, "
        "specifying the main context, key characters, and primary events. The summary must guide the tone and style of the translation "
        "(e.g., somber, emotional). Do not translate individual lines, only provide the summary in Vietnamese.\n\n"
        "Knowledge base (user-provided notes):\n" + get_knowledge_text(load_workbook(XLSX_PATH))
    )
    user_prompt = f"Sheet: {sheet_name}\n\nContent:\n{context}\n\nSummarize in 2-3 lines in Vietnamese."
    try:
        resp = client.responses.create(
            model=os.environ.get("OPENAI_MODEL", "gpt-5-mini"),
            reasoning=Reasoning(effort="medium"),
            instructions=sys_prompt,
            input=user_prompt
        )
        # print(f"Summary Input: ")
        # print(sys_prompt + "\n" + user_prompt)
        return (resp.output_text or "").strip()
    except Exception as e:
        print(f"Error generating summary for {sheet_name}: {e}")
        return ""

def translate_ai(num_lines: int) -> None:
    if not is_file_editable(XLSX_PATH):
        print(f"Excel sheet {XLSX_PATH} is not editable. Skipping.")
        sys.exit(1)

    if not os.path.exists(XLSX_PATH):
        print(f"translate.xlsx not found at {XLSX_PATH}. Run parse first.")
        sys.exit(1)

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("Environment variable OPENAI_API_KEY is not set.")
        sys.exit(1)

    client = OpenAI(api_key=api_key)
    model = os.environ.get("OPENAI_MODEL", "gpt-5-mini")

    wb = load_workbook(XLSX_PATH)
    knowledge_text = get_knowledge_text(wb)
    processed = 0

    sum_ws = wb[SUMMARIES_SHEETNAME] if SUMMARIES_SHEETNAME in wb.sheetnames else None

    for sheet_name in wb.sheetnames:
        if sheet_name in SPECIAL_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [(ws.cell(row=1, column=c).value or "").strip() for c in range(1, 10)]
        try:
            col_id = headers.index("ID") + 1
            col_orig = headers.index("Original") + 1
            col_chinese = headers.index("Chinese") + 1
            col_mtl = headers.index("MTL") + 1
            col_edited = headers.index("Edited") + 1
        except ValueError:
            print(f"Warning: Sheet {sheet_name} has invalid headers. Skipping.")
            continue

        # Thu thập các dòng cần dịch
        rows_to_translate = []
        row_indices = []
        for r in range(2, ws.max_row + 1):
            if processed >= num_lines:
                break
            row_id = (ws.cell(row=r, column=col_id).value or "").strip()
            if not row_id:
                continue
            mtl_val = (ws.cell(row=r, column=col_mtl).value or "").strip()
            if mtl_val:
                continue
            original = ws.cell(row=r, column=col_orig).value or ""
            chinese = ws.cell(row=r, column=col_chinese).value or ""
            if not original and not chinese:
                continue
            rows_to_translate.append((row_id, original, chinese))
            row_indices.append(r)
            processed += 1

        if not rows_to_translate:
            continue

        # Tạo tóm tắt nếu cần
        summary = ""
        if sum_ws:
            for row in sum_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] == sheet_name:
                    summary = row[1] or ""
                    break
        if not summary:
            summary = generate_file_summary(client, sheet_name, rows_to_translate)
            if summary and sum_ws:
                sum_ws.append([sheet_name, summary])
                print(f"Summary for {sheet_name}: {summary}")

        # Tạo prompt duy nhất cho toàn bộ sheet
        prompt_lines = []
        for idx, (row_id, original, chinese) in enumerate(rows_to_translate, 1):
            prompt_lines.append(
                f"Line {idx}:\n"
                f"ID: {row_id}\n"
                f"Original value (source 1): {original or '<empty>'}\n"
                f"Chinese value (source 2): {chinese or '<empty>'}\n"
            )
        content = "\n".join(prompt_lines)
        sys_prompt = (
            f"Knowledge base (user-provided notes):\n{knowledge_text or '<empty>'}\n\n"
            f"File summary:\n{summary or '<no summary>'}\n\n"
            "You are a translator for a visual novel. Translate the following lines into Vietnamese. "
            "Return the translations in a numbered list corresponding to each line's index. "
            "Preserve placeholders, variables, control codes, line breaks, speaker tone, honorifics where appropriate, and context. "
            "Do not provide explanations, only the translations in the format:\n"
            "1. <translation>\n2. <translation>\n..."
        )
        user_prompt = f"Sheet: {sheet_name}\n\nContent:\n{content}\n\nTranslate into Vietnamese as a numbered list."

        try:
            resp = client.responses.create(
                model=model,
                reasoning=Reasoning(effort="medium"),
                instructions=sys_prompt,
                input=user_prompt
            )
            # print(f"Translate Input: ")
            # print(sys_prompt + "\n" + user_prompt)
            ai_text = (resp.output_text or "").strip()
            if not ai_text:
                print(f"Warning: Empty response for {sheet_name}")
                continue

            # print(ai_text)

            # Phân tích phản hồi thành danh sách các bản dịch
            translations = []
            current_translation = []
            current_num = None
            for line in ai_text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                match = re.match(r'^(\d+)\.\s*(.*)$', line)
                if match:
                    if current_translation and current_num is not None:
                        translations.append((current_num, '\n'.join(current_translation).strip()))
                    current_num = int(match.group(1))
                    current_translation = [match.group(2).strip()]
                else:
                    if current_num is not None:
                        current_translation.append(line)
            if current_translation and current_num is not None:
                translations.append((current_num, '\n'.join(current_translation).strip()))

            # Gán bản dịch vào các ô tương ứng
            for num, translation in translations:
                if num > len(rows_to_translate):
                    print(f"Warning: Translation index {num} exceeds number of rows in {sheet_name}")
                    continue
                row_idx = row_indices[num - 1]
                if translation.lower().startswith("tóm tắt") or "summary" in translation.lower():
                    print(f"Warning: AI output for {sheet_name} | Line {num} contains summary: {translation}")
                    continue
                if translation:
                    ws.cell(row=row_idx, column=col_mtl).value = translation
                    print(f"Translated: {sheet_name} | ID {rows_to_translate[num - 1][0]}. Result: {translation}")

        except Exception as e:
            print(f"OpenAI API error on {sheet_name}: {e}")

        if processed >= num_lines:
            break

    if processed > 0 or (sum_ws and sum_ws.max_row > 1):
        update_overview(wb)
        wb.save(XLSX_PATH)
        print(f"Saved {processed} AI translations and summaries to {XLSX_PATH}")
    else:
        print("No rows required translation or already filled.")

def _list_bundles(folder_path: str) -> List[Path]:
    """Return filtered list of bundle paths under folder_path, excluding ignored suffixes."""
    folder = Path(folder_path)
    bundle_paths = list(folder.rglob("*.bundle"))
    return [p for p in bundle_paths if not any(p.name.endswith(suf) for suf in IGNORED_BUNDLE_SUFFIXES)]

def _asset_filename(obj) -> str|None:
    if not obj.container: return None
    return obj.container.split('/')[-1]

def unpack_bundle(folder_path: str) -> None:
    bundle_paths = _list_bundles(folder_path)
    if not bundle_paths:
        print(f"No .bundle files found in {folder_path}")
        return

    os.makedirs(ORIGINAL_DIR, exist_ok=True)

    print(f"Found {len(bundle_paths)} .bundle files to unpack:")

    for bundle_path in bundle_paths:
        try:
            bundle = UnityPy.load(str(bundle_path))
            for obj in bundle.objects:
                if obj.type.name == "TextAsset":
                    data = obj.read()
                    file_name = _asset_filename(obj)
                    out_path = os.path.join(ORIGINAL_DIR, file_name)
                    os.makedirs(os.path.dirname(out_path), exist_ok=True)
                    with open(out_path, "w", encoding="utf-8", newline="") as f:
                        f.write(data.m_Script)
                    print(f"Extracted {file_name} from {bundle_path}")
        except Exception as e:
            print(f"Error unpacking {bundle_path}: {e}")

def rebuild_translated_files() -> None:
    if not os.path.exists(XLSX_PATH):
        print(f"translate.xlsx not found at {XLSX_PATH}. Run parse first.")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)
    if METADATA_SHEETNAME not in wb.sheetnames:
        print("Metadata sheet not found in translate.xlsx")
        sys.exit(1)

    meta_ws = wb[METADATA_SHEETNAME]
    mappings: List[Tuple[str, str, int]] = []
    for r in meta_ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v is None for v in r):
            continue
        sheet_name, mapped_file, ftype = r[:3]
        if not sheet_name or not mapped_file or not ftype:
            continue
        mappings.append((str(sheet_name), str(mapped_file), int(ftype)))

    if os.path.exists(TRANSLATED_DIR):
        shutil.rmtree(TRANSLATED_DIR)
    os.makedirs(TRANSLATED_DIR, exist_ok=True)

    for sheet_name, mapped_file, ftype in mappings:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found. Skipping {mapped_file}.")
            continue
        ws = wb[sheet_name]
        id_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            _id = (row[0] or "").strip()
            if not _id:
                continue
            original = (row[1] or "")
            chinese = trim_blank_lines(row[2] or "")
            mtl = trim_blank_lines(row[3] or "")
            edited = trim_blank_lines(row[4] or "")
            id_rows.append((_id, original, chinese, mtl, edited))

        out_path = os.path.join(TRANSLATED_DIR, mapped_file)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)

        lines_out: List[str] = []

        def add_comment_block(original_text: str, chinese_text: str, mtl_text: str):
            if original_text.strip() != "":
                for ln in (original_text or "").replace('\r\n', '\n').replace('\r', '\n').split('\n'):
                    lines_out.append("; " + ln if ln.strip() != "" else ";")
            if chinese_text.strip() != "":
                lines_out.append("; **Chinese**")
                for ln in (chinese_text or "").replace('\r\n', '\n').replace('\r', '\n').split('\n'):
                    lines_out.append("; " + ln if ln.strip() != "" else ";")
            if mtl_text.strip() != "":
                lines_out.append("; **MTL**")
                for ln in (mtl_text or "").replace('\r\n', '\n').replace('\r', '\n').split('\n'):
                    lines_out.append("; " + ln if ln.strip() != "" else ";")

        if ftype == 2:
            for _id, original, chinese, mtl, edited in id_rows:
                used_value = edited.strip() if edited.strip() != "" else mtl.strip() if mtl.strip() != "" else chinese
                used_value = trim_blank_lines(used_value)
                add_comment_block(original, chinese, mtl)
                lines_out.append(f"{_id}: {used_value}")
                lines_out.append("")
        elif ftype == 1:
            for _id, original, chinese, mtl, edited in id_rows:
                used_value = edited.strip() if edited.strip() != "" else mtl.strip() if mtl.strip() != "" else chinese
                used_value = trim_blank_lines(used_value)
                lines_out.append(f"# {_id}")
                add_comment_block(original, chinese, mtl)
                if used_value == "":
                    lines_out.append("")
                else:
                    for ln in used_value.split('\n'):
                        lines_out.append(ln)
                lines_out.append("")
        else:
            print(f"Warning: Unknown file type {ftype} for {mapped_file}. Skipping.")
            continue

        with open(out_path, 'w', encoding='utf-8', newline='') as f:
            f.write("\n".join(lines_out) + "\n")
        print(f"Wrote {out_path}")

def pack_translated_files(folder_path: str) -> None:
    folder = Path(folder_path)
    bundle_paths = _list_bundles(folder_path)

    if not bundle_paths:
        print(f"No .bundle files found in {folder_path}")
        return

    backup_folder = Path(folder_path + "_backup")
    backup_folder.mkdir(parents=True, exist_ok=True)
    print(f"Using backup folder: {backup_folder}")

    print(f"Found {len(bundle_paths)} .bundle files:")

    translated_text_file_dict = {}
    for file in Path(TRANSLATED_DIR).rglob("*.txt"):
        translated_text_file_dict[file.name] = file

    patched_asset_file_dict = {}
    for ext in ["png", "jpg"]:
        for file in Path(PATCHES_DIR).rglob(f"*.{ext}"):
            patched_asset_file_dict[file.stem] = file

    # Load patches once
    patches = load_patches_from_files()
    # Build a global set of all patch entries to track unpatched across bundles
    all_patch_entries = set()
    if patches:
        for _suf, _id_map in patches.items():
            for _pid, _entries in _id_map.items():
                for _ent in _entries:
                    _sel = _ent.get('object_selector')
                    if _sel is not None:
                        all_patch_entries.add((_suf, _pid, _sel))

    for bundle_path in bundle_paths:
        try:
            bundle_path_str = str(bundle_path)
            bundle = UnityPy.load(bundle_path_str)
            bundle_modified = False

            # Determine relevant patch keys (suffixes) for this bundle
            applicable_suffixes = [suf for suf in patches.keys() if bundle_path_str.endswith(suf)] if patches else []
            patched_count = 0

            opened_objects_by_type = {"Texture2D": []}

            for obj in bundle.objects:
                if obj.type.name == "Texture2D":
                    opened_objects_by_type["Texture2D"].append(obj.read())

            for obj in bundle.objects:
                if obj.type.name == "TextAsset":
                    name = _asset_filename(obj)
                    if name not in translated_text_file_dict:
                        continue

                    with open(translated_text_file_dict[name], 'r', encoding='utf-8') as f:
                        translated_text = f.read()

                    data = obj.read()
                    data.m_Script = translated_text
                    data.save()
                    bundle_modified = True
                    print(f"    Replaced {name} in {bundle_path_str}")
                elif obj.type.name == "SpriteAtlas":
                    data = obj.read()
                    # Find Texture2D, whose name includes data.name
                    matching_texture = None
                    for tex in opened_objects_by_type["Texture2D"]:
                        if data.m_Name in tex.m_Name:
                            matching_texture = tex
                    if matching_texture is None:
                        print(f"Warning: Could not find texture for {data.m_Name} in {bundle_path_str}")
                        continue
                    atlas_image = matching_texture.image  # PIL Image
                    # Ensure atlas is in RGBA for alpha compositing
                    if atlas_image.mode != "RGBA":
                        atlas_image = atlas_image.convert("RGBA")
                    for idx, sprite_name in enumerate(data.m_PackedSpriteNamesToIndex):
                        if sprite_name in patched_asset_file_dict:
                            sprite_image = Image.open(patched_asset_file_dict[sprite_name]).convert("RGBA")
                            coords = data.m_RenderDataMap[idx][1].textureRect
                            x0 = round(coords.x)
                            y0 = round(atlas_image.height - coords.y - sprite_image.height)
                            # 1) Erase destination area first (make fully transparent)
                            transparent_patch = Image.new("RGBA", (sprite_image.width, sprite_image.height), (0, 0, 0, 0))
                            atlas_image.paste(transparent_patch, (x0, y0))
                            # 2) Alpha composite the sprite over the atlas at the destination
                            atlas_image.alpha_composite(sprite_image, dest=(x0, y0))
                            bundle_modified = True
                            print(f"    Patched sprite {sprite_name} in Texture2D {matching_texture.m_Name} in {bundle_path_str}")
                            patched_count += 1
                    if bundle_modified:
                        matching_texture.image = atlas_image
                        matching_texture.save()

                elif obj.type.name == "Texture2D":
                    data = obj.read()
                    if data.m_Name not in patched_asset_file_dict:
                        continue
                    sprite_image = Image.open(patched_asset_file_dict[data.m_Name])
                    data.image = sprite_image
                    data.save()
                    bundle_modified = True
                    print(f"    Patched Texture2D {data.m_Name} in {bundle_path_str}")
                    patched_count += 1

                elif obj.type.name == "MonoBehaviour" and applicable_suffixes:
                    pid_key = str(obj.path_id)
                    todo_entries = []  # list of (suffix, entry)
                    for suf in applicable_suffixes:
                        id_map = patches.get(suf, {})
                        for _ent in id_map.get(pid_key, []):
                            todo_entries.append((suf, _ent))
                    if not todo_entries:
                        continue

                    try:
                        tree = obj.read_typetree()
                    except Exception as e:
                        print(f"Failed to parse MonoBehaviour {bundle_path.name}. Error: {e}")
                        continue
                    # Apply patches
                    any_patched_this_obj = False
                    for suf, ent in todo_entries:
                        selector = ent.get('object_selector')
                        value = ent.get('patched_value')
                        if selector is None:
                            continue
                        ok = set_by_selector(tree, selector, value)
                        if ok:
                            any_patched_this_obj = True
                            patched_count += 1
                            all_patch_entries.discard((suf, pid_key, selector))

                    if any_patched_this_obj:
                        try:
                            obj.save_typetree(tree)
                            bundle_modified = True
                        except Exception as e:
                            print(f"Failed to save typetree for {bundle_path.name} pid {pid_key}: {e}")

            if bundle_modified:
                try:
                    rel_path = bundle_path.relative_to(folder)
                except ValueError:
                    rel_path = Path(bundle_path.name)
                backup_path = backup_folder / rel_path
                backup_path.parent.mkdir(parents=True, exist_ok=True)
                if not backup_path.exists():
                    shutil.copy2(bundle_path, backup_path)
                    print(f"Backed up original: {backup_path}")

                bundle.save(pack="lz4", out_path=os.path.dirname(bundle_path))
                if patched_count > 0:
                    print(f"    Patched {patched_count} in {bundle_path.name}")
                print(f"Saved {bundle_path_str}")

        except Exception as e:
            print(f"Error processing {bundle_path}: {e} \n{traceback.print_stack()}")

    # Global report of unpatched patch entries across all bundles
    print(f"Unpatched entries across all bundles: {len(all_patch_entries)}")
    for suf, pid, sel in sorted(all_patch_entries):
        print(f" - Suffix {suf} | PathID {pid} | selector {sel}")

def refresh():
    if not os.path.exists(XLSX_PATH):
        print(f"translate.xlsx not found at {XLSX_PATH}.")
        sys.exit(1)

    if not is_file_editable(XLSX_PATH):
        print(f"Excel sheet {XLSX_PATH} is not editable. Skipping.")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)

    # Check new .txt files in ORIGINAL_DIR
    existing_sheets = set(wb.sheetnames)
    existing_meta = set()
    if METADATA_SHEETNAME in wb.sheetnames:
        for r in wb[METADATA_SHEETNAME].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                existing_meta.add(str(r[0]))

    new_metadata_rows = []
    for fname in sorted(os.listdir(ORIGINAL_DIR)):
        if not fname.lower().endswith(".txt"):
            continue
        base_sheet_name = os.path.splitext(fname)[0]
        sheet_name = sanitize_sheet_name(base_sheet_name)

        if sheet_name in existing_sheets or sheet_name in existing_meta:
            continue  # already exists

        path = os.path.join(ORIGINAL_DIR, fname)
        ftype, data = _load_and_parse_original_txt(path)
        if ftype is None:
            print(f"Warning: Could not detect file type for {fname}. Skipping.")
            continue

        final_sheet_name = _add_sheet_with_parsed_data(wb, base_sheet_name, data)
        new_metadata_rows.append([final_sheet_name, fname, ftype])
        print(f"Added new sheet for {fname} -> {final_sheet_name}")

    # Update metadata sheet
    if new_metadata_rows:
        if METADATA_SHEETNAME not in wb.sheetnames:
            meta_ws = wb.create_sheet(title=METADATA_SHEETNAME)
            meta_ws.append(METADATA_HEADER)
        else:
            meta_ws = wb[METADATA_SHEETNAME]
        for row in new_metadata_rows:
            meta_ws.append(row)

    # Ensure Patch addresses sheet exists and populate from file if empty
    ensure_patch_sheet(wb)

    # Reupdate overview sheet
    update_overview(wb)
    wb.save(XLSX_PATH)
    print(f"Refreshed {XLSX_PATH}. Added {len(new_metadata_rows)} new sheet(s).")


# --- Binary patching support ---
# Hard-coded mapping of source hex -> destination hex. Ensure same length for each pair.
# Use uppercase or lowercase hex; spaces are ignored. Example entry shown below.
BIN_PATCH_MAP = {
    "0c 00 00 00 e7 ae 80 e4 bd 93 e4 b8 ad e6 96 87 07 00 00 00 7a 68 2d 48 61 6e 74 00 0c 00 00 00 e7 b9 81 e9 ab 94 e4 b8 ad e6 96 87"
    : "0E 00 00 00 54 69 E1 BA BF 6E 67 20 56 69 E1 BB 87 74 00 00 07 00 00 00 7a 68 2d 48 61 6e 74 00 06 00 00 00 e9 ab 94 e4 b8 ad 00 00"
}


def perform_binary_patch(file_path: str, mapping: dict | None = None):
    """Perform hex-based binary patching on a file.

    Steps:
    - Validate mapping pairs have equal length.
    - Read file bytes.
    - Replace all occurrences of each source pattern with its destination bytes.
    - Create a .bak backup before writing any changes.
    """
    if mapping is None:
        mapping = BIN_PATCH_MAP

    if not os.path.isfile(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)

    byte_map = validate_bin_patch_map(mapping)

    with open(file_path, 'rb') as f:
        data = f.read()

    total_replacements = 0
    details: list[tuple[str, str, int]] = []

    for src_bytes, dst_bytes in byte_map.items():
        count = data.count(src_bytes)
        if count > 0:
            data = data.replace(src_bytes, dst_bytes)
            details.append((src_bytes.hex().upper(), dst_bytes.hex().upper(), count))
            total_replacements += count

    if total_replacements == 0:
        print("Bin patch was not performed; no changes made.")
        return

    backup_path = file_path + ".bak"
    if not os.path.exists(backup_path):
        shutil.copy2(file_path, backup_path)
        print(f"Backup created: {backup_path}")

    with open(file_path, 'wb') as f:
        f.write(data)

    print(f"Patched {file_path}: {total_replacements} replacement(s).")
    for s_hex, d_hex, c in details:
        print(f" - {c}x {s_hex} -> {d_hex}")

def gui():
    try:
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import filedialog
        from io import StringIO
    except Exception as e:
        print("customtkinter is required for the GUI. Please install it with: pip install customtkinter")
        print(f"Import error: {e}")
        sys.exit(1)

    # Configure appearance
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("dark-blue")

    def is_valid_exe(path: str) -> bool:
        if not path:
            return False
        try:
            # Check if the file exists and is named manosaba.exe (case-insensitive)
            if not (os.path.isfile(path) and os.path.basename(path).lower() == "manosaba.exe"):
                return False
            # Check for game directory structure to verify legitimate copy
            exe_dir = os.path.dirname(path)
            required_paths = [
                os.path.join(exe_dir, "manosaba_Data"),
                os.path.join(exe_dir, "manosaba_Data", "StreamingAssets"),
                os.path.join(exe_dir, "manosaba_Data", "resources.assets")
            ]
            return all(os.path.exists(p) for p in required_paths)
        except Exception:
            return False

    class TextboxStream(StringIO):
        def __init__(self, textbox):
            super().__init__()
            self.textbox = textbox

        def write(self, text):
            self.textbox.configure(state="normal")
            self.textbox.insert(tk.END, text)
            self.textbox.see(tk.END)  # Auto-scroll to the bottom
            self.textbox.configure(state="disabled")
            self.textbox.update()  # Ensure immediate update

        def flush(self):
            pass

    class App(ctk.CTk):
        def __init__(self):
            super().__init__()
            self.title("Magical Girl Witch Trials Patcher")
            self.geometry("700x400")  # Increased height for log box
            self.resizable(False, False)

            # Grid config
            self.grid_columnconfigure(0, weight=1)
            self.grid_rowconfigure(4, weight=1)  # Make log row expandable

            # Title/Instruction
            self.label_info = ctk.CTkLabel(self, text="Hãy tìm đến file manosaba.exe trong máy và nhấn Patch")
            self.label_info.grid(row=0, column=0, columnspan=3, padx=12, pady=(16, 8), sticky="w")

            # Path entry + Browse
            self.entry_path = ctk.CTkEntry(self, placeholder_text="Đường dẫn manosaba.exe")
            self.entry_path.grid(row=1, column=0, columnspan=2, padx=(12, 6), pady=6, sticky="ew")
            self.entry_path.bind("<KeyRelease>", self.on_path_change)

            self.path = get_steam_game_path('manosaba_game/manosaba.exe')
            if self.path and is_valid_exe(self.path):
                self.entry_path.insert(0, self.path)

            self.btn_browse = ctk.CTkButton(self, text="Browse", command=self.on_browse)
            self.btn_browse.grid(row=1, column=2, padx=(6, 12), pady=6)

            # Patch button
            self.btn_patch = ctk.CTkButton(self, text="Patch", state="disabled", command=self.do_patch)
            self.btn_patch.grid(row=2, column=0, padx=12, pady=(6, 6), sticky="w")

            # Status label
            self.status_var = tk.StringVar(value="")
            self.label_status = ctk.CTkLabel(self, textvariable=self.status_var)
            self.label_status.grid(row=3, column=0, columnspan=3, padx=12, pady=(6, 6), sticky="w")

            # Log text box
            self.log_textbox = ctk.CTkTextbox(self, height=100, state="disabled")
            self.log_textbox.grid(row=4, column=0, columnspan=3, padx=12, pady=(6, 12), sticky="nsew")

            # Redirect print output to log text box
            self.log_stream = TextboxStream(self.log_textbox)
            sys.stdout = self.log_stream

        def set_status(self, message: str, status: str | None = None):
            self.status_var.set(message)
            if status == 'success':
                self.label_status.configure(text_color='#17e841')
            elif status == 'error':
                self.label_status.configure(text_color='#e82817')
            else:
                self.label_status.configure(text_color='#ffffff')

        def on_path_change(self, _):
            self.path = self.entry_path.get().strip()
            if is_valid_exe(self.path):
                self.btn_patch.configure(state="normal")
                self.set_status("Đường dẫn hợp lệ, sẵn sàng patch!", 'success')
            else:
                self.btn_patch.configure(state="disabled")
                self.set_status("Patch chỉ hỗ trợ bản quyền", 'error')

        def on_browse(self):
            file_path = filedialog.askopenfilename(
                title="Mở manosaba.exe",
                filetypes=[("Executable", "*.exe")]
            )
            if file_path:
                self.entry_path.delete(0, tk.END)
                self.entry_path.insert(0, file_path)
                self.on_path_change(None)

        def do_patch(self):
            if not is_valid_exe(self.path):
                self.set_status("Patch chỉ hỗ trợ bản quyền", 'error')
                return
            exe_dir = os.path.dirname(self.path)
            aa_dir = os.path.join(exe_dir, "manosaba_Data", "StreamingAssets", "aa", "StandaloneWindows64")
            res_assets = os.path.join(exe_dir, "manosaba_Data", "resources.assets")
            try:
                self.set_status("Đang patch game...")
                self.log_textbox.configure(state="normal")
                self.log_textbox.delete("1.0", tk.END)  # Clear previous logs
                self.log_textbox.configure(state="disabled")
                print("Bắt đầu quá trình patch...")
                print(f"Đường dẫn game: {self.path}")
                print(f"Đang xử lý thư mục StreamingAssets: {aa_dir}")
                pack_translated_files(aa_dir)
                print(f"Đang patch file resources.assets: {res_assets}")
                perform_binary_patch(res_assets)
                print("Hoàn tất quá trình patch!")
                self.set_status("Xong! Các file đã được patch thành công!", 'success')
            except Exception as e:
                print(f"Lỗi: {str(e)}")
                self.set_status(f"Đã có lỗi xảy ra: {e}", 'error')

        def destroy(self):
            # Restore stdout before closing
            sys.stdout = sys.__stdout__
            super().destroy()

    app = App()
    app.mainloop()

def main():
    command_usage = "python translate_tool.py [unpack <folder>|parse|refresh|translate <num>|build|pack <folder>|build+pack <folder>|binpatch <file>]"
    if len(sys.argv) < 2:
        if is_running_in_exe():
            gui()
            sys.exit(0)
        else:
            print(f"Usage: {command_usage}")
            sys.exit(1)

    cmd = sys.argv[1].lower()
    if cmd == 'parse':
        parse_original_files()
    elif cmd == 'unpack' and len(sys.argv) >= 3:
        unpack_bundle(sys.argv[2])
    elif cmd == 'build':
        rebuild_translated_files()
        dump_patches_from_files()
    elif cmd == 'pack' and len(sys.argv) >= 3:
        pack_translated_files(sys.argv[2])
    elif cmd == 'build+pack' and len(sys.argv) >= 3:
        rebuild_translated_files()
        dump_patches_from_files()
        pack_translated_files(sys.argv[2])
    elif cmd == 'translate' and len(sys.argv) >= 3:
        try:
            n = int(sys.argv[2])
        except ValueError:
            print("For 'translate', provide a number of lines to process. Example: translate 10")
            sys.exit(1)
        if n <= 0:
            print("Number of lines must be positive.")
            sys.exit(1)
        translate_ai(n)
    elif cmd == 'refresh':
        refresh()
    elif cmd == 'binpatch' and len(sys.argv) >= 3:
        perform_binary_patch(sys.argv[2])
    elif cmd == 'gui':
        gui()
    else:
        print(f"Unknown command. Use {command_usage}.")
        sys.exit(1)

if __name__ == '__main__':
    main()