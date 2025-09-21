import os
import sys
from pathlib import Path
import xml.etree.ElementTree as ET
import UnityPy
from openpyxl import Workbook, load_workbook

from translate_tool import (
    apply_header_and_column_widths,
    apply_wrap_to_all_cells,
    load_patches_from_files,
    ensure_patch_sheet,
    populate_patch_sheet_from_file
)

# Configuration
OUTPUT_XLSX = "bundle_info.xlsx"
ADDRESSES_PATH = os.path.join("patches", "addresses.txt")
IGNORED_BUNDLE_SUFFIXES = ['general-managedtext_assets_all.bundle']
IGNORED_CONTAINERS = ['Assets/#WitchTrials/Data/ScriptableObjects/SpecialThanksData.asset']

# Container lookup configuration
ROOT = os.path.dirname(os.path.abspath(__file__))
CONTAINER_LOOKUP_DIR = os.path.join(ROOT, "container_lookup")

# Excel sheet constants
SHEET_NAME = "Bundle Info"
HEADER = ["Bundle Path Suffix", "Container", "Name", "Type", "PathID", "Original Object Selector", "Original",
          "Chinese Object Selector", "Chinese"]

PATCH_SHEET_NAME = "Patch Addresses"
PATCH_HEADER = ["Bundle path suffix", "PathID", "Object selector", "Original", "Translated", "Notes"]

def extract_localized_texts(tree, base_path="", bundle_suffix="", parent_stack=None):
    """Recursively extract localized texts, pairing locale 0 and 2 for the same item index.
    Additionally, maintain a parent path stack (list) that holds the ancestry while traversing.
    This can be useful for future context-aware processing. The return format remains unchanged.
    """
    if parent_stack is None:
        parent_stack = []
    texts = []
    if isinstance(tree, dict):
        for key, value in tree.items():
            new_path = f"{base_path}.{key}" if base_path else key
            # push current key into parent stack
            parent_stack.append(value)
            if isinstance(value, list) and value:
                if isinstance(value[0], dict) and "_locale" in value[0] and "_text" in value[0]:
                    locale_texts = {}
                    for idx, item in enumerate(value):
                        selector = f"{new_path}[{idx}]._text"
                        locale = item.get("_locale")
                        text = item.get("_text")
                        locale_texts[locale] = (selector, text)
                    orig_selector = locale_texts.get(0, ("", ""))[0]
                    orig_text = locale_texts.get(0, ("", ""))[1]
                    cn_selector = locale_texts.get(2, ("", ""))[0]
                    cn_text = locale_texts.get(2, ("", ""))[1]
                    # If original text is missing, check parent for a _defaultText fallback
                    if (not orig_text) and cn_text and len(parent_stack) >= 2:
                        parent_obj = parent_stack[-2]
                        if isinstance(parent_obj, dict) and "_defaultText" in parent_obj:
                            orig_text = parent_obj.get("_defaultText")
                            orig_selector = f"{base_path}._defaultText" if base_path else "_defaultText"
                    if orig_text or cn_text:
                        texts.append((orig_selector, orig_text, cn_selector, cn_text))
                else:
                    for idx, item in enumerate(value):
                        item_path = f"{new_path}[{idx}]"
                        # push index context
                        parent_stack.append(idx)
                        texts.extend(extract_localized_texts(item, item_path, bundle_suffix, parent_stack))
                        parent_stack.pop()
            else:
                texts.extend(extract_localized_texts(value, new_path, bundle_suffix, parent_stack))
            # pop current key
            parent_stack.pop()
    elif isinstance(tree, list):
        for idx, item in enumerate(tree):
            new_path = f"{base_path}[{idx}]"
            parent_stack.append(idx)
            texts.extend(extract_localized_texts(item, new_path, bundle_suffix, parent_stack))
            parent_stack.pop()
    return texts

def get_extracted_texts(obj, bundle_suffix=""):
    """Extract object selectors and texts for original (locale 0) and Chinese (locale 2)."""
    if obj.type.name == "MonoBehaviour":
        try:
            tree = obj.read_typetree()
            extracted = extract_localized_texts(tree, bundle_suffix=bundle_suffix)
            if extracted:
                return extracted
            # if 'm_text' in tree:
            #     return [("m_text", tree['m_text'], "m_text", tree['m_text'])]
        except:
            return []
    return []

def _load_container_lookup_map(dir_path: str = CONTAINER_LOOKUP_DIR) -> dict:
    """Aggregate PathID->Container mapping from XML files in the given folder.
    The XML schema is:
      <Assets>
        <Asset>
          <Container>...</Container>
          <PathID>...</PathID>
        </Asset>
      </Assets>
    The folder may be empty or not exist.
    Returns a dict mapping str(PathID) -> Container.
    """
    mapping = {}
    if not os.path.isdir(dir_path):
        return mapping
    try:
        for fname in os.listdir(dir_path):
            if not fname.lower().endswith('.xml'):
                continue
            fpath = os.path.join(dir_path, fname)
            try:
                tree = ET.parse(fpath)
                root = tree.getroot()
                for asset in root.findall('./Asset'):
                    pid_el = asset.find('PathID')
                    cont_el = asset.find('Container')
                    pid = pid_el.text.strip() if pid_el is not None and pid_el.text else None
                    cont = cont_el.text.strip() if cont_el is not None and cont_el.text else None
                    if pid and cont and pid not in mapping:
                        mapping[pid] = cont
            except Exception as e:
                print(f"Warning: Failed to parse container lookup file {fpath}: {e}")
    except Exception as e:
        print(f"Warning: Could not read container lookup dir {dir_path}: {e}")
    return mapping

def generate_bundle_info(folder_path: str):
    """Generate an Excel file with bundle asset information, grouping by container."""
    folder = Path(folder_path)
    bundle_paths = [p for p in folder.rglob("*.bundle") if
                    not any(str(p).endswith(suf) for suf in IGNORED_BUNDLE_SUFFIXES)]

    if not bundle_paths:
        print(f"No .bundle files found in {folder_path}")
        return

    # Load patches
    patches = load_patches_from_files()

    # Load container lookup map (optional)
    container_map = _load_container_lookup_map()

    # Create or load Excel workbook (avoid overwriting existing file)
    if os.path.exists(OUTPUT_XLSX):
        try:
            wb = load_workbook(OUTPUT_XLSX)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # Ensure/Create main info sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        # Clear existing data rows, keep header if present
        if ws.max_row and ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        # Ensure header exists and widths applied
        if (ws.cell(row=1, column=1).value or "").strip() != HEADER[0]:
            ws.delete_rows(1, ws.max_row)
            ws.append(HEADER)
            apply_header_and_column_widths(ws, HEADER, [40, 60, 30, 15, 16, 30, 60, 30, 60])
        else:
            apply_header_and_column_widths(ws, HEADER, [40, 60, 30, 15, 16, 30, 60, 30, 60])
    else:
        ws = wb.create_sheet(title=SHEET_NAME)
        ws.append(HEADER)
        apply_header_and_column_widths(ws, HEADER, [40, 60, 30, 15, 16, 30, 60, 30, 60])

    # Collect all asset data grouped by bundle
    bundle_data = {}
    for bundle_path in bundle_paths:
        bundle_suffix = str(bundle_path.relative_to(folder))
        bundle_data[bundle_suffix] = []

        try:
            bundle = UnityPy.load(str(bundle_path))

            for obj in bundle.objects:
                if obj.type.name not in ["MonoBehaviour"]:
                    continue

                resolved_container = obj.container or container_map.get(str(obj.path_id))
                if resolved_container in IGNORED_CONTAINERS:
                    continue

                name = obj.read_typetree()['m_Name']

                extracted = get_extracted_texts(obj, bundle_suffix)
                for orig_selector, original, cn_selector, chinese in extracted:
                    bundle_data[bundle_suffix].append({
                        "container": resolved_container,
                        "name": name,
                        "type": obj.type.name,
                        "path_id": str(obj.path_id),
                        "original_selector": orig_selector,
                        "original": original,
                        "chinese_selector": cn_selector,
                        "chinese": chinese,
                    })

                if obj.type.name == "MonoBehaviour" and bundle_suffix in patches:
                    id_map = patches.get(bundle_suffix, {})
                    pid_str = str(obj.path_id)
                    if pid_str in id_map:
                        for entry in id_map[pid_str]:
                            selector = entry.get('object_selector', '')
                            if selector:
                                patched_entry = {
                                    "container": resolved_container,
                                    "name": name,
                                    "type": obj.type.name,
                                    "path_id": pid_str,
                                    "original_selector": selector,
                                    "original": entry.get('patched_value', ''),
                                    "chinese_selector": selector,
                                    "chinese": entry.get('patched_value', ''),
                                }
                                bundle_data[bundle_suffix].append(patched_entry)

            print(f"Processed {bundle_path}")

        except Exception as e:
            print(f"Error processing {bundle_path}: {e}")

    # Write to Excel, including all fields for every row
    all_rows_for_patch = []
    for bundle_suffix in sorted(bundle_data.keys()):
        assets = bundle_data[bundle_suffix]
        if not assets:
            continue

        sorted_assets = sorted(assets, key=lambda x: (x["container"] or "", x["name"], x["type"], x["path_id"]))

        for asset in sorted_assets:
            ws.append([
                bundle_suffix,
                asset["container"],
                asset["name"],
                asset["type"],
                asset["path_id"],
                asset["original_selector"],
                asset["original"],
                asset["chinese_selector"],
                asset["chinese"]
            ])

            # Build Notes: Name, Container, and Original value (with line break after ':') if they exist
            notes_lines = []
            if asset.get("name"):
                notes_lines.append(f"Name: {asset['name']}")
            if asset.get("container"):
                notes_lines.append(f"Container: {asset['container']}")
            if asset.get("original"):
                notes_lines.append("Original value:\n" + str(asset["original"]))
            notes_text = "\n".join(notes_lines)

            all_rows_for_patch.append([
                bundle_suffix,
                asset["path_id"],
                asset["chinese_selector"],
                asset["chinese"],
                asset.get("translated", ""),
                notes_text
            ])

    apply_wrap_to_all_cells(ws)

    # Merge into Patch Addresses sheet without overwriting existing data
    ws_patch = ensure_patch_sheet(wb)
    # Map headers and build index
    headers = [(ws_patch.cell(row=1, column=c).value or "").strip() for c in range(1, ws_patch.max_column + 1)]
    try:
        col_suffix = headers.index("Bundle path suffix") + 1
        col_pathid = headers.index("PathID") + 1
        col_selector = headers.index("Object selector") + 1
        col_original = headers.index("Original") + 1
        col_translated = headers.index("Translated") + 1
        col_notes = headers.index("Notes") + 1
    except ValueError:
        # Recreate header if mismatched
        ws_patch.delete_rows(1, ws_patch.max_row)
        ws_patch.append(PATCH_HEADER)
        apply_header_and_column_widths(ws_patch, PATCH_HEADER, [40, 16, 40, 60, 60, 60])
        col_suffix, col_pathid, col_selector, col_original, col_translated, col_notes = 1, 2, 3, 4, 5, 6

    index = {}
    for r in range(2, ws_patch.max_row + 1):
        suf = (ws_patch.cell(row=r, column=col_suffix).value or "").strip()
        pid = str((ws_patch.cell(row=r, column=col_pathid).value or "").strip())
        sel = (ws_patch.cell(row=r, column=col_selector).value or "").strip()
        if suf and pid and sel:
            index[(suf, pid, sel)] = r

    # Add/merge rows from this run
    for row in all_rows_for_patch:
        suf, pid, sel, original, translated, notes = row
        key = (str(suf).strip(), str(pid).strip(), str(sel).strip())
        if not key[0] or not key[1] or not key[2]:
            continue
        r = index.get(key)
        if r is None:
            ws_patch.append([key[0], key[1], key[2], original, translated, notes])
            index[key] = ws_patch.max_row
        else:
            # Fill Original/Notes if empty; leave Translated to user/patch
            o_cell = ws_patch.cell(row=r, column=col_original)
            if (o_cell.value is None) or (str(o_cell.value).strip() == ""):
                o_cell.value = original
            n_cell = ws_patch.cell(row=r, column=col_notes)
            if (n_cell.value is None) or (str(n_cell.value).strip() == ""):
                n_cell.value = notes

    # Now append any missing data from files without overwriting existing rows
    populate_patch_sheet_from_file(wb, update_instead_of_overwrite=True)

    apply_wrap_to_all_cells(ws_patch)

    wb.save(OUTPUT_XLSX)
    print(f"Saved bundle information to {OUTPUT_XLSX}")

def main():
    command_usage = "python bundle_info.py [info <folder>]"
    if len(sys.argv) < 3:
        print(f"Usage: {command_usage}")
        sys.exit(1)

    cmd = sys.argv[1].lower()
    folder_path = sys.argv[2]

    if not os.path.isdir(folder_path):
        print(f"Error: {folder_path} is not a valid directory")
        sys.exit(1)

    if cmd == 'info':
        generate_bundle_info(folder_path)
    else:
        print(f"Unknown command. Use {command_usage}.")
        sys.exit(1)

if __name__ == "__main__":
    main()
